from urllib import request

from fastapi import (
    FastAPI,
    Request,
    Query,
    UploadFile,
    File,
    HTTPException
)
from fastapi.responses import (
    HTMLResponse,
    FileResponse,
    RedirectResponse
)
from starlette.middleware.sessions import SessionMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

import openpyxl
import joblib
import pandas as pd
from datetime import datetime, timedelta
from app.mail import send_otp, generate_otp


from app.database import (
    create_table,
    add_product,
    get_low_stock_products,
    get_products,
    update_product,
    delete_product,
    get_dashboard_stats,
    get_category_data,
    get_stock_data,
    get_inventory_value_data,
    get_all_categories,
    get_total_categories,
    get_connection,
    register_user,
    login_user,
    update_password
)

from app.model import (
    Product,
    Prediction,
    UserRegister,
    UserLogin
)

app = FastAPI(title="Smart Inventory Optimization System")
otp_storage = {}
app.add_middleware(
    SessionMiddleware,
    secret_key="smart_inventory_secret_2026"
)


# -----------------------------------
# Static Files
# -----------------------------------

app.mount(
    "/static",
    StaticFiles(directory="static"),
    name="static"
)

templates = Jinja2Templates(directory="templates")


# -----------------------------------
# Load ML Model
# -----------------------------------

model = joblib.load("models/demand_model.pkl")
encoder = joblib.load("models/category_encoder.pkl")


# -----------------------------------
# Startup
# -----------------------------------

@app.on_event("startup")
def startup():
    create_table()

# -----------------------------------
# Check Login
# -----------------------------------


def check_login(request: Request):
    if "user" not in request.session:
        raise HTTPException(
            status_code=401,
            detail="Please Login First"
        )



# -----------------------------------
# HTML Pages
# -----------------------------------

@app.get("/", response_class=HTMLResponse)
def login_page(request: Request):

    if "user" in request.session:
        return RedirectResponse("/dashboard")

    return templates.TemplateResponse(
        request=request,
        name="login.html"
    )


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard_page(request: Request):

    if "user" not in request.session:
        return RedirectResponse("/", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="dashboard.html",
        context={
            "user": request.session["user"]
        }
    )


@app.get("/products", response_class=HTMLResponse)
def products_page(request: Request):

    if "user" not in request.session:
        return RedirectResponse("/", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="products.html",
        context={
            "user": request.session["user"]
        }
    )


@app.get("/analytics", response_class=HTMLResponse)
def analytics_page(request: Request):

    if "user" not in request.session:
        return RedirectResponse("/", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="analytics.html",
        context={
            "user": request.session["user"]
        }
    )


@app.get("/prediction", response_class=HTMLResponse)
def prediction_page(request: Request):

    if "user" not in request.session:
        return RedirectResponse("/", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="prediction.html",
        context={
            "user": request.session["user"]
        }
    )

# -----------------------------------
# Dashboard API
# -----------------------------------

@app.get("/api/dashboard")
def dashboard_data(request: Request):

    check_login(request)

    stats = get_dashboard_stats()
    stats["total_categories"] = get_total_categories()

    return stats

# -----------------------------------
# Analytics API
# -----------------------------------

@app.get("/api/analytics")
def analytics_data(request: Request):

    check_login(request)

    dashboard = get_dashboard_stats()
    dashboard["total_categories"] = get_total_categories()

    return {
        "dashboard": dashboard,
        "category_data": get_category_data(),
        "stock_data": get_stock_data(),
        "inventory_value_data": get_inventory_value_data()
    }

# -----------------------------------
# Prediction API
# -----------------------------------

@app.post("/api/predict")
def predict_inventory(request: Request, data: Prediction):

    check_login(request)

    category = encoder.transform([data.category])[0]

    input_data = pd.DataFrame([{
        "category": category,
        "price": data.price,
        "stock": data.stock,
        "daily_sales": data.daily_sales,
        "lead_time": data.lead_time
    }])

    predicted = model.predict(input_data)[0]

    recommended_stock = round(predicted)

    restock_status = (
        "Required"
        if data.stock < recommended_stock
        else "Not Required"
    )

    if predicted >= 100:
        demand_level = "High"
    elif predicted >= 40:
        demand_level = "Medium"
    else:
        demand_level = "Low"

    return {
        "predicted_demand": round(predicted),
        "recommended_stock": recommended_stock,
        "restock_status": restock_status,
        "demand_level": demand_level
    }

# -----------------------------------
# Product APIs
# -----------------------------------

@app.post("/api/products")
def create_product(request: Request, product: Product):

    check_login(request)

    success = add_product(product)

    if not success:
        raise HTTPException(
            status_code=409,
            detail="Product Already Exists"
        )

    return {
        "message": "Product Added Successfully"
    }
    

# Get Products (Pagination + Search + Category Filter)
@app.get("/api/products")
def all_products(
    request: Request,
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=100, ge=1, le=500),
    search: str = "",
    category: str = "All",
    stock_filter: str = "All",
    sort_by: str = "id",
    sort_order: str = "desc"
):

    check_login(request)

    return get_products(
        page=page,
        limit=limit,
        search=search,
        category=category,
        stock_filter=stock_filter,
        sort_by=sort_by,
        sort_order=sort_order
    )

# Update Product
@app.put("/api/products/{id}")
def edit_product(request: Request, id: int, product: Product):

    check_login(request)

    success = update_product(id, product)

    if not success:
        raise HTTPException(
            status_code=409,
            detail="Product Already Exists"
        )

    return {
        "message": "Product Updated Successfully"
    }

# Delete Product
@app.delete("/api/products/{id}")
def remove_product(request: Request, id: int):

    check_login(request)

    delete_product(id)

    return {
        "message": "Product Deleted Successfully"
    }

@app.get("/api/categories")
def categories(request: Request):

    check_login(request)

    return get_all_categories()
# -----------------------------------
# Low Stock Products API
# -----------------------------------

@app.get("/api/low-stock")
def low_stock_products(request: Request):

    check_login(request)

    return get_low_stock_products()

# -----------------------------------
# Export Products to Excel
# -----------------------------------

@app.get("/api/export")
def export_products(request: Request):

    check_login(request)

    products = get_products(
        page=1,
        limit=100000,
        search="",
        category="All",
        sort_by="id",
        sort_order="desc"
    )["products"]

    
    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet.title = "Products"

    sheet.append([
        "ID",
        "Product Name",
        "Category",
        "Stock",
        "Price"
    ])

    for product in products:

        sheet.append([
            product["id"],
            product["product_name"],
            product["category"],
            product["stock"],
            product["price"]
        ])

    filename = "Smart_Inventory_Products.xlsx"

    workbook.save(filename)

    return FileResponse(
        filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="Smart_Inventory_Products.xlsx"
    )
# -----------------------------------
# Import Products from Excel
# -----------------------------------


@app.post("/api/import")
async def import_products(
    request: Request,
    file: UploadFile = File(...)
):

    check_login(request)


    # Only Excel file allowed
    if not file.filename.lower().endswith(".xlsx"):
        return {
            "message": "Please upload Excel (.xlsx) file",
            "imported": 0,
            "skipped": 0
        }

    conn = None

    try:

        workbook = openpyxl.load_workbook(file.file)
        sheet = workbook.active

        conn = get_connection()
        cursor = conn.cursor()

        imported = 0
        skipped = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):

            # Skip empty rows
            if (
                not row or
                row[1] is None or
                row[2] is None or
                row[3] is None or
                row[4] is None
            ):
                continue

            # Check duplicate
            cursor.execute(
                """
                SELECT id
                FROM products
                WHERE LOWER(product_name)=LOWER(?)
                AND LOWER(category)=LOWER(?)
                """,
                (
                    str(row[1]).strip(),
                    str(row[2]).strip()
                )
            )

            if cursor.fetchone():
                skipped += 1
                continue

            # Insert Product
            cursor.execute(
                """
                INSERT INTO products
                (product_name, category, stock, price)
                VALUES (?, ?, ?, ?)
                """,
                (
                    str(row[1]).strip(),
                    str(row[2]).strip(),
                    int(row[3]),
                    float(row[4])
                )
            )

            imported += 1

        conn.commit()

        return {
            "message": "Import Completed Successfully",
            "imported": imported,
            "skipped": skipped
        }

    except Exception as e:

        print(e)

        return {
            "message": "Import Failed",
            "imported": 0,
            "skipped": 0
        }

    finally:

        if conn:
            conn.close()
# -----------------------------------
# Register API
# -----------------------------------

@app.post("/api/register")
def register(request: Request, data: UserRegister):

    success = register_user(data)

    if not success:
        raise HTTPException(
            status_code=409,
            detail="Email already registered"
        )

    request.session["user"] = {
        "name": data.name,
        "email": data.email
    }

    return {
        "message": "Registration Successful"
    }


# -----------------------------------
# Login API
# -----------------------------------

@app.post("/api/login")
def login(request: Request, data: UserLogin):

    user = login_user(data)

    if user is None:
        raise HTTPException(
            status_code=401,
            detail="Invalid Email or Password"
        )

    request.session["user"] = {
        "id": user["id"],
        "name": user["name"],
        "email": user["email"]
    }

    return {
        "message": "Login Successful"
    }


# -----------------------------------
# Reset Password API
# -----------------------------------

@app.post("/api/reset-password")
def reset_password(data: dict):

    email = data.get("email")
    new_password = data.get("password")

    if not email or not new_password:
        raise HTTPException(
            status_code=400,
            detail="Email and Password Required"
        )

    success = update_password(email, new_password)

    if not success:
        raise HTTPException(
            status_code=404,
            detail="User Not Found"
        )

    if email in otp_storage:
        del otp_storage[email]

    return {
        "message": "Password Updated Successfully"
    }

@app.get("/register", response_class=HTMLResponse)
def register_page(request: Request):

    if "user" in request.session:
        return RedirectResponse("/dashboard", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="register.html"
    )


@app.get("/api/user")
def current_user(request: Request):

    check_login(request)

    return request.session["user"]


@app.get("/logout")
def logout(request: Request):

    request.session.clear()

    response = RedirectResponse(url="/", status_code=303)

    response.delete_cookie("session")

    return response


@app.post("/api/send-otp")
def send_reset_otp(data: dict):

    email = data.get("email")

    if not email:
        raise HTTPException(status_code=400, detail="Email Required")

    otp = generate_otp()

    otp_storage[email] = {
        "otp": otp,
        "expiry": datetime.now() + timedelta(minutes=5)
    }

    send_otp(email, otp)

    return {
        "message": "OTP Sent Successfully"
    }


@app.post("/api/verify-otp")
def verify_reset_otp(data: dict):

    email = data.get("email")
    otp = data.get("otp")

    if email not in otp_storage:
        raise HTTPException(status_code=400, detail="OTP Not Found")

    saved = otp_storage[email]

    if datetime.now() > saved["expiry"]:
        del otp_storage[email]
        raise HTTPException(status_code=400, detail="OTP Expired")

    if otp != saved["otp"]:
        raise HTTPException(status_code=400, detail="Invalid OTP")

    return {
        "message": "OTP Verified Successfully"
    }


@app.get("/forgot-password", response_class=HTMLResponse)
def forgot_password_page(request: Request):

    return templates.TemplateResponse(
        request=request,
        name="forgot_password.html"
    )

